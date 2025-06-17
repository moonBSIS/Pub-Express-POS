from fastapi import FastAPI, Depends, HTTPException, Body, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import func, update
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from database import *
from schemas import *
from typing import List, Optional
from datetime import date as dt_date, datetime, time
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import csv, io, shutil, os, openpyxl

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ✅ Create the tables
Base.metadata.create_all(bind=engine)

# ✅ Dependency to get the DB session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/")
def read_root():
    return {"message": "Hello, FastAPI!"}


# --- Admin Sales Report --- #


@app.get("/summary/range", response_model=SalesSummary)
def get_sales_summary_range(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db),
):
    try:
        start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d"), time.min)
        end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d"), time.max)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    # --- Paid Orders ---
    orders_paid = db.query(Order).filter(
        Order.is_paid == True,
        Order.created_at >= start_dt,
        Order.created_at <= end_dt
    ).all()

    # --- Unpaid Orders ---
    orders_unpaid = db.query(Order).filter(
        Order.is_paid == False,
        Order.created_at >= start_dt,
        Order.created_at <= end_dt
    ).all()

    # --- Total Sales ---
    total_paid = sum(order.total for order in orders_paid)
    total_unpaid = sum(order.total for order in orders_unpaid)

    # --- Dine-in ---
    dine_in_orders = [o for o in orders_paid if (o.type or "").lower() == "dine-in"]
    dine_in_count = len(dine_in_orders)
    dine_in_sales = sum(o.total for o in dine_in_orders)

    # --- Take-out ---
    take_out_orders = [o for o in orders_paid if (o.type or "").lower() == "take-out"]
    take_out_count = len(take_out_orders)
    take_out_sales = sum(o.total for o in take_out_orders)
    
    # --- Void Orders ---
    void_orders = db.query(Order).filter(
        Order.is_paid == True,
        Order.cancel_reason.isnot(None),
        Order.created_at >= start_dt,
        Order.created_at <= end_dt
    ).count()

    return {
        "total_paid": total_paid,
        "total_unpaid": total_unpaid,
        "orders_paid": len(orders_paid),
        "orders_unpaid": len(orders_unpaid),
        "total_discount": sum(order.discount or 0 for order in orders_paid),
        "dine_in_count": dine_in_count,
        "dine_in_sales": dine_in_sales,
        "take_out_count": take_out_count,
        "take_out_sales": take_out_sales,
        "void_orders": void_orders
        
    }


# --- Database Backup --- #


@app.on_event("startup")
def auto_backup_monthly():
    backup_folder = "backups"
    os.makedirs(backup_folder, exist_ok=True)

    today = datetime.now()
    backup_filename = f"backup_{today.year}-{today.month:02d}.db"
    backup_path = os.path.join(backup_folder, backup_filename)

    if not os.path.exists(backup_path):
        shutil.copyfile("pub_express.db", backup_path)
        print(f"Monthly backup created: {backup_filename}")
    else:
        print(f"Monthly backup already exists for {today.year}-{today.month:02d}")


# --- Dashboard --- #


@app.get("/dashboard/today-summary")
def get_today_summary(db: Session = Depends(get_db)):
    today = dt_date.today()
    start = datetime.combine(today, datetime.min.time())
    end = datetime.combine(today, datetime.max.time())

    total_sales = db.query(func.sum(Order.total)).filter(
        Order.is_paid == True,
        Order.created_at.between(start, end),
        Order.cancel_reason == None
    ).scalar() or 0.0

    total_discount = db.query(func.sum(Order.discount)).filter(
        Order.is_paid == True,
        Order.created_at.between(start, end),
        Order.cancel_reason == None
    ).scalar() or 0.0

    total_orders = db.query(Order).filter(
        Order.is_paid == True,
        Order.created_at.between(start, end),
        Order.cancel_reason == None
    ).count()

    # New: Count void orders
    void_orders = db.query(Order).filter(
        Order.cancel_reason.isnot(None),
        Order.created_at.between(start, end)
    ).count()

    return {
        "total_sales": round(total_sales, 2),
        "total_discount": round(total_discount, 2),
        "total_orders": total_orders,
        "average_order_value": round(total_sales / total_orders, 2) if total_orders else 0.0,
        "void_orders": void_orders
    }

@app.get("/reports/top-items")
def top_selling_items(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    limit: int = 10,
    db: Session = Depends(get_db),
):
    query = db.query(
        MenuItem.name,
        func.sum(OrderItem.quantity).label("total_sold")
    ).join(OrderItem.menu_item).join(OrderItem.order)

    if start_date and end_date:
        start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d"), time.min)
        end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d"), time.max)

        query = query.filter(
            Order.is_paid == True,
            Order.created_at >= start_dt,
            Order.created_at <= end_dt
        )
    else:
        today = dt_date.today()
        start_dt = datetime.combine(today, datetime.min.time())
        end_dt = datetime.combine(today, datetime.max.time())
        query = query.filter(
            Order.is_paid == True,
            Order.created_at >= start_dt,
            Order.created_at <= end_dt
        )

    results = query.group_by(MenuItem.name).order_by(
        func.sum(OrderItem.quantity).desc()
    ).limit(limit).all()

    return [{"item": name, "total_sold": total_sold} for name, total_sold in results]


@app.get("/reports/item-sales")
def item_sales(
    date: str = Query(...),
    item_id: int = Query(...),
    db: Session = Depends(get_db),
):
    try:
        date_dt = datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format.")

    start = datetime.combine(date_dt, datetime.min.time())
    end = datetime.combine(date_dt, datetime.max.time())

    total_quantity = db.query(func.sum(OrderItem.quantity)) \
        .join(OrderItem.order) \
        .filter(
            Order.is_paid == True,
            OrderItem.menu_item_id == item_id,
            Order.created_at >= start,
            Order.created_at <= end
        ) \
        .scalar() or 0

    return {
        "date": date,
        "total_sales": total_quantity
    }


def top_selling_items(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    limit: int = 10,
    db: Session = Depends(get_db),
):
    query = db.query(
        MenuItem.name,
        func.sum(OrderItem.quantity).label("total_sold")
    ).join(OrderItem.menu_item).join(OrderItem.order)

    if start_date and end_date:
        start_dt = datetime.combine(datetime.strptime(start_date, "%Y-%m-%d"), time.min)
        end_dt = datetime.combine(datetime.strptime(end_date, "%Y-%m-%d"), time.max)

        query = query.filter(
            Order.is_paid == True,
            Order.created_at >= start_dt,
            Order.created_at <= end_dt
        )
    else:
        # Default to today's date if not provided
        today = dt_date.today()
        start_dt = datetime.combine(today, datetime.min.time())
        end_dt = datetime.combine(today, datetime.max.time())
        query = query.filter(
            Order.is_paid == True,
            Order.created_at >= start_dt,
            Order.created_at <= end_dt
        )

    results = query.group_by(MenuItem.name).order_by(
        func.sum(OrderItem.quantity).desc()
    ).limit(limit).all()

    return [{"item": name, "total_sold": total_sold} for name, total_sold in results]

@app.get("/reports/orders-xlsx")
def export_orders_xlsx(start_date: str, end_date: str, db: Session = Depends(get_db)):
    orders = (
        db.query(Order)
        .filter(Order.created_at >= start_date)
        .filter(Order.created_at <= end_date + " 23:59:59")
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"

    headers = [
        "Date", "Order #", "Order Type", "Order Item", "Comments",
        "Quantity Sold", "Discount Type", "Discount Name", "Discount ID",
        "Total Discount", "Total Amount", "Net Total", "Cashier"
    ]
    ws.append(headers)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="df4444", end_color="df4444", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    item_totals = {}

    for order in orders:
        total = float(order.total or 0)
        discount = float(order.discount or 0)
        net_total = total - discount

        first_row = True
        for item in order.items:
            name = item.menu_item.name
            item_totals[name] = item_totals.get(name, 0) + item.quantity

            discount_type = item.discount_person_type or ""
            discount_name = item.discount_person_name or ""
            discount_id = item.discount_person_id or ""

            # For coupons/vouchers, reflect type as "Coupon", use code as ID
            if discount_type.lower() in ["coupon", "voucher"]:
                discount_type = "Coupon"
                discount_id = item.discount_person_id or "Code"

            row = [
                order.created_at.strftime("%Y-%m-%d %H:%M") if first_row else "",
                order.id if first_row else "",
                order.type or "N/A" if first_row else "",
                name,
                item.notes or "",
                item.quantity,
                discount_type if first_row else "",
                discount_name if first_row else "",
                discount_id if first_row else "",
                discount if first_row else "",
                total if first_row else "",
                net_total if first_row else "",
                order.cashier or "" if first_row else "",
            ]
            ws.append(row)
            first_row = False

    # Auto width for Sheet 1
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

    # === Sheet 2: Item Totals ===
    ws2 = wb.create_sheet("Item Totals")
    ws2.append(["Item", "Total Quantity Sold"])

    for item_name, total_qty in item_totals.items():
        ws2.append([item_name, total_qty])

    for col_num in range(1, 3):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1dbfc1", end_color="1dbfc1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for column in ws2.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws2.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PubExpress_Sales_Report_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/dashboard/discount-usage")
def get_discount_usage(db: Session = Depends(get_db)):
    today = dt_date.today()
    start = datetime.combine(today, datetime.min.time())
    end = datetime.combine(today, datetime.max.time())

    results = (
        db.query(OrderItem.discount_person_type, func.count(OrderItem.id))
        .join(OrderItem.order)
        .filter(Order.created_at.between(start, end))
        .filter(Order.is_paid == True)
        .filter(OrderItem.discount_person_type.isnot(None))
        .group_by(OrderItem.discount_person_type)
        .all()
    )

    return [{"type": row[0], "count": row[1]} for row in results]

# --- Sales Report --- #

@app.get("/reports/daily-sales")
def daily_sales_report(date: Optional[dt_date] = Query(default=None), db: Session = Depends(get_db)):
    if date is None:
        date = dt_date.today()

    start_datetime = datetime.combine(date, datetime.min.time())
    end_datetime = datetime.combine(date, datetime.max.time())

    total_sales = db.query(func.sum(Order.total)) \
                    .filter(Order.is_paid == True) \
                    .filter(Order.created_at >= start_datetime) \
                    .filter(Order.created_at <= end_datetime) \
                    .scalar() or 0.0

    total_orders = db.query(func.count(Order.id)) \
                     .filter(Order.is_paid == True) \
                     .filter(Order.created_at >= start_datetime) \
                     .filter(Order.created_at <= end_datetime) \
                     .scalar()

    return {
        "date": date.isoformat(),
        "total_orders": total_orders,
        "total_sales": total_sales
    }


@app.get("/reports/discount-usage")
def discount_usage_summary(
    db: Session = Depends(get_db),
):
    results = (
        db.query(
            OrderItem.discount_person_type,
            func.count(OrderItem.id).label("count")
        )
        .filter(OrderItem.discount_person_type.isnot(None))
        .group_by(OrderItem.discount_person_type)
        .all()
    )
    return [{"type": d_type, "count": count} for d_type, count in results]

@app.get("/summary/today", response_model=SalesSummary)
def get_today_summary(db: Session = Depends(get_db)):
    today = dt_date.today()
    start = datetime.combine(today, datetime.min.time())
    end = datetime.combine(today, datetime.max.time())
    return compute_sales_summary(start, end, db)


def compute_sales_summary(start_dt: datetime, end_dt: datetime, db: Session):
    orders_paid = db.query(Order).filter(
        Order.is_paid == True,
        Order.created_at >= start_dt,
        Order.created_at <= end_dt
    ).all()

    orders_unpaid = db.query(Order).filter(
        Order.is_paid == False,
        Order.created_at >= start_dt,
        Order.created_at <= end_dt
    ).all()

    total_paid = sum(order.total for order in orders_paid)
    total_unpaid = sum(order.total for order in orders_unpaid)

    return {
        "total_paid": total_paid,
        "total_unpaid": total_unpaid,
        "orders_paid": len(orders_paid),
        "orders_unpaid": len(orders_unpaid)
    }

# --- Menu Endpoints --- #

@app.post("/menu/", response_model=MenuItemOut)
def create_menu_item(item: MenuItemCreate, db: Session = Depends(get_db)):
    db_item = MenuItem(**item.dict())
    db.add(db_item)
    try:
        db.commit()
        db.refresh(db_item)
        return db_item
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Menu item name already exists.")

@app.get("/menu/", response_model=list[MenuItemOut])
def get_all_menu_items(db: Session = Depends(get_db)):
    return db.query(MenuItem).all()

@app.put("/menu/{item_id}", response_model=MenuItemOut)
def update_menu_item(
    item_id: int,
    item_update: MenuItemUpdate,
    db: Session = Depends(get_db),
):
    item = db.query(MenuItem).filter(MenuItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    item.name = item_update.name
    item.category = item_update.category
    item.price = item_update.price

    db.commit()
    db.refresh(item)
    return item

@app.put("/orders/{order_id}/discount")
def update_order_discount(
    order_id: int,
    discount_data: DiscountUpdate,
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.discount = discount_data.discount
    db.commit()
    return {"message": f"Discount updated to {discount_data.discount} for order #{order_id}"}

@app.delete("/menu/{item_id}")
def delete_menu_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(MenuItem).get(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"message": "Item deleted"}

# --- Order Endpoints --- #

@app.post("/orders/", response_model=OrderOut)
def create_order(
    order_data: OrderCreate,
    db: Session = Depends(get_db),
):
    total = 0
    order = Order(
        notes=order_data.notes,
        type=order_data.type,
        discount=order_data.discount or 0.0,
        cashier=order_data.cashier,
    )
    db.add(order)
    db.flush()  # flush so order.id is available

    for item in order_data.items:
        menu_item = db.query(MenuItem).filter(MenuItem.id == item.menu_item_id).first()
        if not menu_item:
            raise HTTPException(status_code=404, detail=f"Menu item {item.menu_item_id} not found")

        # Calculate subtotal
        subtotal = menu_item.price * item.quantity
        total += subtotal

        # Save OrderItem with possible discount person details
        db_item = OrderItem(
            order_id=order.id,
            menu_item_id=menu_item.id,
            quantity=item.quantity,
            discount_person_name=item.discount_person_name,
            discount_person_id=item.discount_person_id,
            discount_person_type=item.discount_person_type,
            manual_discount_type=item.manual_discount_type,
            manual_discount_value=item.manual_discount_value,
            notes=item.notes,
        )
        db.add(db_item)

    # Set final total and payment status
    order.total = total
    due = total - (order_data.discount or 0.0)
    order.paid = order_data.paid_amount or 0.0
    order.is_paid = order.paid >= due

    db.commit()
    db.refresh(order)
    return order


@app.post("/orders/new", response_model=OrderOut)
def create_empty_order(
    db: Session = Depends(get_db),
):
    new_order = Order(total=0.0, is_paid=False, paid=0.0)
    db.add(new_order)
    db.commit()
    db.refresh(new_order)
    return new_order

@app.post("/orders/{order_id}/items", response_model=OrderItemOut)
def add_item_to_order(
    order_id: int,
    item_data: OrderItemCreate,
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    menu_item = db.query(MenuItem).filter(MenuItem.id == item_data.menu_item_id).first()
    if not menu_item:
        raise HTTPException(status_code=404, detail="Menu item not found")

    order_item = OrderItem(
        order_id=order_id,
        menu_item_id=item_data.menu_item_id,
        quantity=item_data.quantity,
    )
    db.add(order_item)

    order.total += menu_item.price * item_data.quantity
    db.commit()
    db.refresh(order_item)
    return order_item

@app.put("/orders/{order_id}/items/{item_id}")
def update_order_item_quantity(
    order_id: int,
    item_id: int,
    update: UpdateQuantity,
    db: Session = Depends(get_db),
):
    order_item = db.query(OrderItem).filter_by(id=item_id, order_id=order_id).first()
    if not order_item:
        raise HTTPException(status_code=404, detail="Order item not found")

    order_item.quantity = update.quantity

    # 🧠 Recalculate order total
    order = db.query(Order).filter(Order.id == order_id).first()
    new_total = 0
    for item in order.items:
        new_total += item.menu_item.price * item.quantity

    order.total = new_total
    db.commit()
    return {"message": "Quantity updated and total recalculated"}

@app.delete("/orders/{order_id}")
def delete_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    db.delete(order)
    db.commit()
    return {"message": f"Order #{order_id} deleted"}

@app.delete("/orders/{order_id}/items/{item_id}")
def delete_order_item(
    order_id: int,
    item_id: int,
    db: Session = Depends(get_db),
):
    order_item = db.query(OrderItem).filter(
        OrderItem.id == item_id, OrderItem.order_id == order_id
    ).first()
    if not order_item:
        raise HTTPException(status_code=404, detail="Order item not found")

    menu_item = db.query(MenuItem).filter(MenuItem.id == order_item.menu_item_id).first()
    order = db.query(Order).filter(Order.id == order_id).first()

    order.total -= menu_item.price * order_item.quantity
    db.delete(order_item)

    db.commit()
    return {"message": "Item removed from order"}

@app.get("/orders/", response_model=List[OrderOut])
def get_orders(
    status: Optional[str] = Query(None),
    is_paid: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.menu_item)
    )

    if status:
        query = query.filter(Order.status == status)
    if is_paid is not None:
        query = query.filter(Order.is_paid == is_paid)

    if not status:
        # Show all orders including canceled unless filtered by status
        pass
    elif status == "void":
        query = query.filter(Order.status == "void")
    else:
        query = query.filter(Order.status == status)
    return query.all()

@app.get("/orders/{order_id}", response_model=OrderOut)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.menu_item)
    ).filter(Order.id == order_id).first()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order

@app.get("/orders/{order_id}/items", response_model=List[OrderItemOut])
def get_order_items(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return order.items

@app.post("/orders/{order_id}/save-unpaid")
def save_unpaid(order_id: int, paid: float = Body(...), db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.paid and order.paid >= order.total:
        raise HTTPException(status_code=400, detail="Order is already paid")

    order.paid = paid
    db.commit()
    return {"message": "Order saved as unpaid", "order_id": order.id, "paid": order.paid}

@app.post("/orders/{order_id}/pay")
def mark_order_paid(
    order_id: int,
    pay_data: PayData = Body(...),
    db: Session = Depends(get_db),
):
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        print(f"Processing payment for Order #{order_id}")
        print(f"Order: {order}")
        print(f"Pay Data: {pay_data}")

        order.paid = pay_data.paid
        order.is_paid = True
        order.status = "paid"

        # === SAFER RECEIPT NUMBERING ===
        if not order.receipt_number:
            year = datetime.now().year
            prefix = f"PX-{year}-"

            # Get the highest existing receipt number that matches the year
            latest_receipt = (
                db.query(Order.receipt_number)
                .filter(Order.receipt_number.like(f"{prefix}%"))
                .order_by(Order.receipt_number.desc())
                .first()
            )

            if latest_receipt and latest_receipt[0]:
                latest_number = int(latest_receipt[0].split("-")[-1])
                next_number = latest_number + 1
            else:
                next_number = 1

            order.receipt_number = f"{prefix}{str(next_number).zfill(6)}"

        db.commit()
        return {
            "message": "Order marked as paid",
            "paid": order.paid,
            "receipt_number": order.receipt_number
        }

    except Exception as e:
        print(f"❌ Error processing payment for Order #{order_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to mark order as paid")

    

@app.delete("/orders/{order_id}/cancel")
def cancel_order(
    order_id: int,
    cancel_data: CancelOrderRequest = Body(...),
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.cancel_reason = cancel_data.cancel_reason or "Canceled by user"

    # ✅ Reset payment info
    order.status = "void"
    order.is_paid = False
    order.paid = 0.0

    db.commit()
    return {
        "message": f"Order #{order_id} canceled",
        "cancel_reason": order.cancel_reason,
        "status": order.status,
    }

@app.post("/orders/hold")
def hold_order(
    order: OrderCreate,
    db: Session = Depends(get_db),
):
    db_order = Order(
        type=order.type,
        notes=order.notes,
        status="held",
        cashier=order.cashier 
    )
    db.add(db_order)
    db.flush() 

    for item in order.items:
        db_item = OrderItem(
            order_id=db_order.id,
            menu_item_id=item.menu_item_id,
            quantity=item.quantity
        )
        db.add(db_item)

    db.commit()
    db.refresh(db_order)
    return {"message": "Order held successfully", "order_id": db_order.id}

@app.post("/orders/{order_id}/restore")
def restore_order(
    order_id: int,
    db: Session = Depends(get_db),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if not order.cancel_reason:
        raise HTTPException(status_code=400, detail="Order is not canceled")

    order.cancel_reason = None
    db.commit()
    return {"message": f"Order #{order_id} restored"}

@app.on_event("startup")
def cleanup_orphan_order_items():
    db = SessionLocal()
    try:
        broken_items = db.query(OrderItem).filter(
            ~OrderItem.menu_item_id.in_(
                db.query(MenuItem.id)
            )
        ).all()

        if broken_items:
            count = len(broken_items)
            print(f"Cleaning up {count} broken OrderItem(s)...")

            for item in broken_items:
                db.delete(item)

            db.commit()
            print(f"Cleanup complete.")

        else:
            print("No broken OrderItems found. Database is clean.")

    finally:
        db.close()

@app.post("/menu/delete-category")
def delete_category(
    category: str = Body(...),
    reassign_to: Optional[str] = Body(default=None),
    db: Session = Depends(get_db),
):
    items_to_update = db.query(MenuItem).filter(MenuItem.category == category).all()

    if reassign_to:
        for item in items_to_update:
            item.category = reassign_to
    else:
        for item in items_to_update:
            db.delete(item)

    db.commit()
    return {"message": f"Category '{category}' deleted and items reassigned." if reassign_to else f"Category '{category}' and its items deleted."}

# --- Supervisor Management Endpoints --- #

@app.get("/supervisors")
def get_supervisors(db: Session = Depends(get_db)):
    return db.query(Supervisor).all()


@app.put("/supervisors/{id}")
def update_supervisor(id: int, data: SupervisorUpdate, db: Session = Depends(get_db)):
    sup = db.query(Supervisor).filter(Supervisor.id == id).first()
    if not sup:
        raise HTTPException(status_code=404, detail="Supervisor not found")
    sup.pin = data.pin
    db.commit()
    return {"message": "Supervisor PIN updated"}


# --- Cashier Management Endpoints --- #


@app.get("/cashiers", response_model=List[CashierOut])
def get_cashiers(db: Session = Depends(get_db)):
    return db.query(Cashier).all()


@app.post("/cashiers", response_model=CashierOut)
def create_cashier(data: CashierCreate, db: Session = Depends(get_db)):
    cashier = Cashier(name=data.name, pin=data.pin)
    db.add(cashier)
    db.commit()
    db.refresh(cashier)
    return cashier


@app.put("/cashiers/{id}")
def update_cashier(id: int, data: CashierUpdate, db: Session = Depends(get_db)):
    cashier = db.query(Cashier).filter(Cashier.id == id).first()
    if not cashier:
        raise HTTPException(status_code=404, detail="Cashier not found")
    if data.name:
        cashier.name = data.name
    if data.pin:
        cashier.pin = data.pin
    db.commit()
    return {"message": "Cashier updated"}


@app.delete("/cashiers/{id}")
def delete_cashier(id: int, db: Session = Depends(get_db)):
    cashier = db.query(Cashier).filter(Cashier.id == id).first()
    if not cashier:
        raise HTTPException(status_code=404, detail="Cashier not found")
    db.delete(cashier)
    db.commit()
    return {"message": "Cashier deleted"}

@app.on_event("startup")
def ensure_supervisor_pin_exists():
    db = SessionLocal()
    try:
        existing = db.query(Supervisor).first()
        if not existing:
            db.add(Supervisor(pin="1234"))  # Default PIN
            db.commit()
            print("Default supervisor PIN (1234) created.")
    finally:
        db.close()

@app.get("/cashiers/{id}")
def get_cashier(id: int, db: Session = Depends(get_db)):
    cashier = db.query(Cashier).filter(Cashier.id == id).first()
    if not cashier:
        raise HTTPException(status_code=404, detail="Cashier not found")
    return cashier


# --- Dynamic QR code generation --- #


@app.get("/settings/qr-link")
def get_qr_link(db: Session = Depends(get_db)):
    setting = db.query(Setting).filter(Setting.key == "qr_link").first()
    return {"value": setting.value if setting else ""}


@app.put("/settings/qr-link")
def update_qr_link(
    data: SettingUpdate,
    db: Session = Depends(get_db),
):
    setting = db.query(Setting).filter(Setting.key == "qr_link").first()
    if setting:
        setting.value = data.value
    else:
        setting = Setting(key="qr_link", value=data.value)
        db.add(setting)

    db.commit()
    return {"message": "QR link updated", "value": setting.value}

# Dynamic QR code message
@app.get("/settings/receipt-message")
def get_receipt_message(db: Session = Depends(get_db)):
    setting = db.query(Setting).filter(Setting.key == "receipt_message").first()
    return {"value": setting.value if setting else ""}


@app.put("/settings/receipt-message")
def update_receipt_message(
    data: SettingUpdate,
    db: Session = Depends(get_db),
):
    setting = db.query(Setting).filter(Setting.key == "receipt_message").first()
    if setting:
        setting.value = data.value
    else:
        setting = Setting(key="receipt_message", value=data.value)
        db.add(setting)
    db.commit()
    return {"message": "Receipt message updated", "value": setting.value}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000, log_config=None)

